"""
Download public reference files and write Parquet (+ optional CSV) for this app.

DOL Form 5500 (official):
  https://www.askebsa.dol.gov/FOIA%20Files/<year>/Latest/F_5500_<year>_Latest.zip
  Contains f_<year>_latest.csv — mapped to columns expected by app.tools.dol.

NAICS 2022 (6-digit titles):
  Default CSV is the NAICS 2022 code list hosted by WID Center (widcenter.org),
  which mirrors the standard NAICS hierarchy/titles. For the official Census
  *industry* crosswalk XLSX only, pass --naics-source census-xlsx (fewer 6-digit rows).

Outputs (by default under ./data):
  - dol_form5500.parquet
  - naics_codes.parquet
  - naics_codes.csv  (same rows; convenient for NAICS_DATA_PATH=.csv)

Usage:
  python scripts/fetch_reference_data.py
  python scripts/fetch_reference_data.py --year 2023 --output-dir ./data
  python scripts/fetch_reference_data.py --skip-dol
"""

from __future__ import annotations

import argparse
import io
import logging
import sys
import tempfile
import zipfile
from pathlib import Path

import duckdb
import httpx

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

DOL_ZIP_TEMPLATE = (
    "https://www.askebsa.dol.gov/FOIA%20Files/{year}/Latest/F_5500_{year}_Latest.zip"
)

# 6-digit NAICS titles (CSV). Census does not publish a single stable 6-digit CSV URL;
# this file is widely used for NAICS 2022 reference tables.
NAICS_WIDCENTER_CSV = "https://data.widcenter.org/download/naics2022/naiccode2022.csv"

CENSUS_INDUSTRY_XLSX = (
    "https://www2.census.gov/programs-surveys/demo/guidance/industry-occupation/"
    "2022-Census-Industry-Code-List-with-Crosswalk.xlsx"
)


def _download(url: str, dest: Path, timeout: float = 600.0) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Downloading %s -> %s", url, dest)
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        with client.stream("GET", url) as resp:
            resp.raise_for_status()
            with dest.open("wb") as f:
                for chunk in resp.iter_bytes(1024 * 1024):
                    f.write(chunk)
    logger.info("Saved %s (%d bytes)", dest, dest.stat().st_size)


def _pick_f5500_csv_member(zf: zipfile.ZipFile) -> str:
    for name in zf.namelist():
        lower = name.lower()
        if "layout" in lower:
            continue
        if lower.endswith(".csv") and "5500" in lower:
            return name
    raise FileNotFoundError("No F_5500 CSV found in zip")


def build_dol_parquet(zip_path: Path, parquet_out: Path) -> None:
    with zipfile.ZipFile(zip_path, "r") as zf:
        member = _pick_f5500_csv_member(zf)
        logger.info("Extracting %s from zip", member)
        raw_csv = zf.read(member)

    tmpdir = Path(tempfile.mkdtemp(prefix="dol5500_"))
    csv_path = tmpdir / "f5500.csv"
    csv_path.write_bytes(raw_csv)

    parquet_out.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(":memory:")
    src = str(csv_path.resolve()).replace("'", "''")
    dst = str(parquet_out.resolve()).replace("'", "''")
    # Map DOL public CSV columns -> app schema (see app/services/database.py, app/tools/dol.py)
    con.execute(
        f"""
        COPY (
          SELECT
            TRIM(CAST("ACK_ID" AS VARCHAR)) AS ack_id,
            TRIM(CAST("SPONS_DFE_EIN" AS VARCHAR)) AS ein,
            TRIM(CAST("PLAN_NAME" AS VARCHAR)) AS plan_name,
            TRIM(CAST("SPONSOR_DFE_NAME" AS VARCHAR)) AS sponsor_dfe_name,
            TRIM(CAST("SPONS_DFE_PN" AS VARCHAR)) AS spons_dfe_pn,
            TRIM(COALESCE(
              NULLIF(TRIM(CAST("SPONS_DFE_MAIL_US_STATE" AS VARCHAR)), ''),
              NULLIF(TRIM(CAST("SPONS_DFE_LOC_US_STATE" AS VARCHAR)), '')
            )) AS spons_state,
            TRIM(CAST("BUSINESS_CODE" AS VARCHAR)) AS business_code,
            CAST(NULL AS VARCHAR) AS broker_name,
            CAST(NULL AS VARCHAR) AS broker_ein,
            CAST(NULL AS VARCHAR) AS cpa_name,
            TRIM(CAST("ADMIN_NAME" AS VARCHAR)) AS plan_admin_name,
            TRIM(CAST("ADMIN_SIGNED_NAME" AS VARCHAR)) AS plan_admin_sign_name,
            TRY_CAST("TOT_PARTCP_BOY_CNT" AS INTEGER) AS tot_partcp_boy_cnt
          FROM read_csv_auto('{src}', nullstr='', ignore_errors=true)
        ) TO '{dst}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """
    )
    con.close()
    pq_esc = str(parquet_out.resolve()).replace("'", "''")
    n = duckdb.connect(":memory:").execute(
        f"SELECT COUNT(*) FROM read_parquet('{pq_esc}')"
    ).fetchone()[0]
    logger.info("Wrote %s (%d rows)", parquet_out, n)


def build_naics_from_widcenter_csv(csv_bytes: bytes, parquet_out: Path, csv_out: Path | None) -> None:
    tmp = Path(tempfile.mkdtemp(prefix="naics_")) / "naics_raw.csv"
    tmp.write_bytes(csv_bytes)

    con = duckdb.connect(":memory:")
    src = str(tmp.resolve()).replace("'", "''")
    dst = str(parquet_out.resolve()).replace("'", "''")
    con.execute(
        f"""
        COPY (
          SELECT
            TRIM(CAST(naicscode AS VARCHAR)) AS naics_code,
            TRIM(CAST(naicstitle AS VARCHAR)) AS naics_title,
            TRIM(CAST(naicsdesc AS VARCHAR)) AS naics_description
          FROM read_csv_auto('{src}', header=true, nullstr='')
          WHERE TRIM(CAST(naicslvl AS VARCHAR)) = '6'
            AND LENGTH(TRIM(CAST(naicscode AS VARCHAR))) = 6
        ) TO '{dst}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """
    )
    if csv_out:
        csv_out.parent.mkdir(parents=True, exist_ok=True)
        pq = str(parquet_out.resolve()).replace("'", "''")
        cs = str(csv_out.resolve()).replace("'", "''")
        con.execute(
            f"""
            COPY (SELECT * FROM read_parquet('{pq}')) TO '{cs}' (HEADER, DELIMITER ',')
            """
        )
    pq_esc = str(parquet_out.resolve()).replace("'", "''")
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{pq_esc}')").fetchone()[0]
    con.close()
    logger.info("Wrote %s (%d 6-digit NAICS rows)", parquet_out, n)
    if csv_out:
        logger.info("Wrote %s", csv_out)


def build_naics_from_census_xlsx(xlsx_path: Path, parquet_out: Path, csv_out: Path | None) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit(
            "census-xlsx NAICS source needs openpyxl: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "2007 to 2022 Industry Crosswalk" not in wb.sheetnames:
        raise SystemExit("Unexpected Census workbook layout (crosswalk sheet missing).")
    ws = wb["2007 to 2022 Industry Crosswalk"]
    # Header row index 6 (0-based): '2022 NAICS code' is column index 13
    seen: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 7:
            continue
        if not row or len(row) < 15:
            continue
        code = row[13]
        desc = row[14]
        if code is None or desc is None:
            continue
        for part in str(code).replace("/", ",").split(","):
            p = part.strip()
            if not p or not p.replace(" ", "").isdigit():
                continue
            digits = p.replace(" ", "")
            if len(digits) == 6:
                seen[digits] = str(desc).strip()

    if not seen:
        raise SystemExit("No 6-digit NAICS rows extracted from Census crosswalk.")

    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE naics (naics_code VARCHAR, naics_title VARCHAR, naics_description VARCHAR)")
    con.executemany(
        "INSERT INTO naics VALUES (?, ?, ?)",
        [(c, t, "") for c, t in sorted(seen.items())],
    )
    parquet_out.parent.mkdir(parents=True, exist_ok=True)
    pq = str(parquet_out.resolve()).replace("'", "''")
    con.execute(f"COPY naics TO '{pq}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    if csv_out:
        cs = str(csv_out.resolve()).replace("'", "''")
        con.execute(f"COPY naics TO '{cs}' (HEADER, DELIMITER ',')")
    logger.info("Wrote %s (%d rows from Census crosswalk)", parquet_out, len(seen))
    con.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Fetch DOL F5500 + NAICS reference data")
    parser.add_argument("--year", type=int, default=2024, help="DOL Form 5500 plan year zip")
    parser.add_argument(
        "--output-dir", type=Path, default=Path("data"), help="Directory for outputs"
    )
    parser.add_argument("--skip-dol", action="store_true")
    parser.add_argument("--skip-naics", action="store_true")
    parser.add_argument(
        "--naics-source",
        choices=("widcenter", "census-xlsx"),
        default="widcenter",
        help="NAICS 6-digit source (widcenter: full 6-digit list; census-xlsx: sparse)",
    )
    parser.add_argument("--keep-zip", action="store_true", help="Keep downloaded DOL zip")
    args = parser.parse_args()
    out = args.output_dir.resolve()
    out.mkdir(parents=True, exist_ok=True)

    if not args.skip_dol:
        year = args.year
        dol_url = DOL_ZIP_TEMPLATE.format(year=year)
        zip_path = out / f"F_5500_{year}_Latest.zip"
        _download(dol_url, zip_path)
        try:
            build_dol_parquet(zip_path, out / "dol_form5500.parquet")
        finally:
            if not args.keep_zip:
                zip_path.unlink(missing_ok=True)

    if not args.skip_naics:
        pq = out / "naics_codes.parquet"
        csv_out = out / "naics_codes.csv"
        if args.naics_source == "widcenter":
            buf = io.BytesIO()
            with httpx.Client(timeout=300.0, follow_redirects=True) as client:
                r = client.get(NAICS_WIDCENTER_CSV)
                r.raise_for_status()
                buf.write(r.content)
            build_naics_from_widcenter_csv(buf.getvalue(), pq, csv_out)
        else:
            xlsx_path = out / "2022-Census-Industry-Code-List-with-Crosswalk.xlsx"
            _download(CENSUS_INDUSTRY_XLSX, xlsx_path)
            try:
                build_naics_from_census_xlsx(xlsx_path, pq, csv_out)
            finally:
                xlsx_path.unlink(missing_ok=True)

    logger.info("Done. Point DOL_DATA_PATH / NAICS_DATA_PATH at the .parquet or .csv files.")


if __name__ == "__main__":
    try:
        main()
    except httpx.HTTPStatusError as e:
        logger.error("HTTP error: %s", e)
        sys.exit(1)
